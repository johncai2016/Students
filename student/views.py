from django.conf import settings
from django.http import HttpResponse
from django.template import loader
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect
from django.utils.crypto import get_random_string
from django.shortcuts import render
import zipfile
import pytesseract
import re
import os
import time
from PIL import Image
import datetime
from django.db import models
from .models import Family,Routing
import threading
from openpyxl import Workbook,load_workbook
import shutil
from django.utils.timezone import utc
from aip import AipOcr
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED, ALL_COMPLETED, as_completed

#这个方法主要是有时候需要调试可以在一个地方设置unique_id,方便统一修改。
def get_unique_id(request):
  unique_id = request.COOKIES.get('unique_id')
  #unique_id = 'qoi7ZLbL'
  return unique_id

def back_to_main_page(msg):
  response = HttpResponse(msg + "<br><a href='/steps'>返回前一页</a>")
  return response

#解压文件到指定目录，因为中文会便乱码，故采用以下解决方案：https://www.jb51.net/article/200832.htm
def unzip_gbk(zip_file_path,unique_id):
    unzip_dir_path = settings.MEDIA_ROOT + unique_id
    with zipfile.ZipFile(file=zip_file_path, mode='r') as zf:
      # 解压到指定目录,首先创建一个解压目录
      os.mkdir(unzip_dir_path)
      for old_name in zf.namelist():
        # 获取文件大小，目的是区分文件夹还是文件，如果是空文件应该不好用。
        file_size = zf.getinfo(old_name).file_size
        # 由于源码遇到中文是cp437方式，所以解码成gbk，windows即可正常
        try:
            new_name = old_name.encode('cp437').decode('gbk')
        except:
            new_name = old_name.encode('utf-8').decode('utf-8')
        # 拼接文件的保存路径
        new_path = os.path.join(unzip_dir_path, new_name)
        #如果文件找不到，说明要先创建目录
        if not os.path.exists(new_path):
           #获取文件的完整的路径用来创建对应的目录
           new_dir = re.search('(.*)/',new_path).group(1)
           #当一个目录下有多个文件时，不需要重复创建目录
           if not os.path.exists(new_dir):
              os.makedirs(new_dir)

        # 如果是文件，则创建对应的文件
        if file_size > 0:
          # 是文件，通过open创建文件，写入数据
          with open(file=new_path, mode='wb') as f:
            # zf.read 是读取压缩包里的文件内容
            f.write(zf.read(old_name))
    
          
@csrf_exempt
def upload_zipFile(request):
    if request.method == "POST":
      myFile = request.FILES.get("myfile")
      if not myFile:
         response = HttpResponse("请上传文件")
         return response
      
      #使用一个8位随机数来保存文件和跟踪是谁上传的文件，这是为了在第二步使用
      unique_id = get_random_string(length=8)
      full_filename = settings.MEDIA_ROOT + unique_id + '.zip'
      destination = open(full_filename,'wb+')
      for chunk in myFile.chunks():
        destination.write(chunk)
      destination.close()
      
      unzip_gbk(full_filename,unique_id)
      
      msg = "上传文件成功!"
      response = back_to_main_page(msg)
      #把这个随机数写入客户的浏览器来确定他的数据。It will expire in 1 hour
      response.set_cookie("unique_id",unique_id,max_age=36000)
    else:
      response = HttpResponse('只允许POST请求.')
    return response
      
@csrf_exempt
def upload_excelFile(request):
    if request.method == "POST":
      myFile = request.FILES.get("myfile")
      if not myFile:
         response = HttpResponse("请上传文件")
         return response
      
      unique_id = request.COOKIES.get('unique_id')
      if unique_id is None:
        response = HttpResponse('请先到Step1上传学生压缩数据再比对.')
      else:
        full_filename = settings.MEDIA_ROOT + unique_id + '.xlsx'
      destination = open(full_filename,'wb+')
      for chunk in myFile.chunks():
        destination.write(chunk)
      destination.close()
      
      msg = "上传文件成功!"
      response = back_to_main_page(msg)
    else:
      response = HttpResponse('只允许POST请求.')
    return response

#首页主要是说明如何使用     
def index(request):
    template = loader.get_template('index.html')
    #这个test参数实际上并没有使用
    context = {
      'test':'test',
    }
    return HttpResponse(template.render(context,request))

#具体实现的功能放在这个页面 
def steps(request):
    template = loader.get_template('steps.html')
    #传当前日期进去
    context = {
      'query_time':datetime.datetime.now().strftime('%Y-%m-%d'),
    }
    return HttpResponse(template.render(context,request))
  
def extract_data_from_img_by_baidu(file_name):
    APP_ID = ''
    API_KEY = ''
    SECRET_KEY = ''
    client = AipOcr(APP_ID, API_KEY, SECRET_KEY)
    img_text = ''
    with open(file_name, 'rb') as fp:
        image = fp.read()
        results = client.general(image)["words_result"]
        #print(results)
        for result in results:
          img_text = img_text + result["words"]
    #print(img_text)
    return img_text

def test(request):
  file_name = '/root/john_project/student/static/0.jpg'
  #text = pytesseract.image_to_string(Image.open(file_name),lang='chi_sim')
  text = extract_data_from_img_by_baidu(file_name)
  response = HttpResponse('从百度提取的文字是:' + text)
  return response


#这个函数用于处理单个学生的数据.
def extract_data_from_img(unique_id,student):
    path = settings.MEDIA_ROOT + unique_id
    image_dir = get_image_dir(settings.MEDIA_ROOT + unique_id)
    
    route_with_star = False
    #存储学生和其家属的名字
    family_name = []
    #存储采样时间
    inspection_time = []
    #存储检测结果
    inspection_result = []
    #存储电话信息
    phone_info = []
    
    #从截图中提取每个学生和家属的数据
    student_path = path + '/' + image_dir + '/' + student      
    files = os.listdir(student_path)
    
    for f in files:
        write_debug_log('正在处理图片文件:' + f)
        image_name = student_path + '/' + f
        text = extract_data_from_img_by_baidu(image_name)
        #text = pytesseract.image_to_string(Image.open(image_name),lang='chi_sim')
        text = re.sub('[\s+]','',text)
        #检查核酸'结果,姓名一般是2-5个中文字
        person_name = re.search("检测中([\u4E00-\u9FA5]{2,5})采样时间",text,re.S)
        collect_time = re.search("采样时间([0-9]{4}-[0-9]{2}-[0-9]{2})?",text)
        result_time = re.search("检测时间([0-9]{4}-[0-9]{2}-[0-9]{2})?",text)
        #结果只会有检测中和阴性，因为如果是阳性，早被拉走了.
        #final_result = re.search("检测结果(.*)",text)
        #提取检测相关信息，如果图片上有采样时间，说明图片是核酸记录
        if collect_time is not None:
           if person_name is not None:
             #目前不知道为啥无法识别‘检测中’的人员姓名
             person_name = str.strip(person_name.group(1))
             
             if person_name == '':
               person_name = '无法识别'
           else:
             person_name = '无法识别'
           
           collect_time = str.strip(collect_time.group(1))
           if result_time is None:
             result_time = '检测中'
             final_result = '检测中'   
           else:
             result_time = str.strip(result_time.group(1))
             final_result = '阴性'
           
           #家庭成员有多人，所以暂时存放在一个list,后续方便遍历存进数据库.
           family_name.append(person_name)
           inspection_time.append(collect_time)
           inspection_result.append(final_result)
           write_debug_log('学生姓名:' + student + '|家庭成员名字:' + person_name + '|检测时间:' + collect_time + '|检测结果:' + final_result)
      
        #检查行程码
        phone = re.search("绿色行程卡(1.*)的动态",text)
        #route_date = re.search("更新于:(.*)",text)
        route_date = datetime.datetime.utcnow()
        with_star_in_result = re.search("中风险",text)
        #目前无法识别行程码时间。
        if phone is not None:
          phone = str.strip(phone.group(1))
          phone_info.append(phone)
          if with_star_in_result is not None:
            route_with_star = True
          write_debug_log('Phone:' + phone)
        
    #把数据临时保存到数据库中以方便在页面按需要展示出来.
    for name,c_time,result in zip(family_name,inspection_time,inspection_result):
        family_obj = Family(unique_id=unique_id,student_name=student,family_name=name,inspection_time=c_time,
                                inspection_result=result)
        family_obj.save()
    
    for phone_num in phone_info:    
        routing_obj = Routing(unique_id=unique_id,student_name=student,phone=phone_num,has_star=route_with_star,
                                show_time=route_date)
        
        routing_obj.save()

#解压缩的时候本身的文件夹也会被提取出来作为一个目录。      
def get_image_dir(path):
  file_list = os.listdir(path)
  for f in file_list:
    if os.path.isdir(path + "/" + f):
       return f

#只需在当前路径创建debug.log就可以实现debug
def write_debug_log(msg):
    debug_file = 'debug.log'
    if os.path.exists(debug_file):
      with open(debug_file,'a') as f:
        f.write(msg + '\r\n')

#处理所有学生数据     
def process_students_data(request):
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
 
  #这是解压后的目录  
  path = settings.MEDIA_ROOT + unique_id
  image_dir = get_image_dir(path)
  write_debug_log('image_dir is:' + image_dir)
  #image_dir_list 才是每个学生的上级目录
  image_dir_list = os.listdir(path + "/" + image_dir)
  start = datetime.datetime.now()
  #f代表每个学生的目录
  for f in image_dir_list:
      extract_data_from_img(unique_id,f)
  #excutor = ThreadPoolExecutor(max_workers=2)
  #all_task = [excutor.submit(extract_data_from_img,(unique_id),(f)) for f in image_dir_list]
  #with ThreadPoolExecutor(max_workers=len(image_dir_list)) as t:
  #    all_task = [t.submit(extract_data_from_img, unique_id, f) for f in image_dir_list]
  #    wait(all_task,return_when=ALL_COMPLETED)
  #wait(all_task,return_when=ALL_COMPLETED)
  #for f in image_dir_list:
  #    extract_data_from_img(unique_id,f)
  end = datetime.datetime.now()
  msg = "数据提取完毕！总处理时间:" + str(end - start)
  response = back_to_main_page(msg)
  return response


#我们只需要查找到一个异常数据就要报告该学生.  
def show_issues(request):
  query_time = request.GET.get('query_time')
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
  
  if query_time is None:
    response = HttpResponse('请加上query_time参数!')
    return response
  #检测中
  no_result_students = Family.objects.filter(unique_id=unique_id,inspection_result='检测中').values('student_name').distinct().order_by('student_name')
  #提交的数据时间不是要求的时间
  time_incorrect_students = Family.objects.filter(unique_id=unique_id,inspection_time__lt=query_time).values('student_name').distinct().order_by('student_name')
  
  # 以下数据来自Routing表
  with_star_students = Routing.objects.filter(unique_id=unique_id,has_star=True).values('student_name').distinct().order_by('student_name')
  # 如果可以提取到时间，也可以像上面一样对时间做检查
  
  template = loader.get_template('show_issues.html')
  context = {
    'no_result_students':no_result_students,
    'time_incorrect_students':time_incorrect_students,
    'with_star_students':with_star_students,
  }
  return HttpResponse(template.render(context,request))


#展示学生和同住人员的核酸详细数据
def show_details(request):
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
  all_students = Family.objects.filter(unique_id=unique_id).order_by('student_name')
  template = loader.get_template('show_all_details.html')
  context = {
    'all_students':all_students,
  }
  return HttpResponse(template.render(context,request))


#显示行程码详细数据
def show_route_details(request):
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
    
  all_routings = Routing.objects.filter(unique_id=unique_id).order_by('student_name')
  template = loader.get_template('show_route_details.html')
  context = {
    'all_routings':all_routings,
  }
  return HttpResponse(template.render(context,request))


#从上传的excel表中获取学生同住人员数量
def load_family_number_from_excel(unique_id):
  path = settings.MEDIA_ROOT + unique_id + '.xlsx'
  #返回期望的学生和同住人员数量
  family_count = {}
  
  #加载excel表，获取学生姓名和同住人员数量。
  excel_file = load_workbook(path)
  all_sheet = excel_file.get_sheet_names()
  #只加载第一张表
  first_sheet = excel_file.get_sheet_by_name(all_sheet[0])
  total_rows = first_sheet.max_row
  #遍历每一行数据(没有表头)
  for row in range(total_rows):
     student_name = first_sheet.cell(row=row + 1,column=1).value
     count = first_sheet.cell(row=row + 1,column=2).value
     family_count[student_name] = count
  return family_count
  
#比对填写的同住人员数量是否匹配提交的核酸数量。
def check_family_number(request):
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
    
  questions_students = []
  expect_counts = []
  actual_counts = []
  family_count = load_family_number_from_excel(unique_id)
  #遍历每个学生然后获取同住人员数量(包含学生)
  for student in family_count.keys():
    expect_count = family_count[student]
    actual_count = Family.objects.filter(unique_id=unique_id,student_name=student).count()
    if expect_count != actual_count:
      questions_students.append(student)
      expect_counts.append(expect_count)
      actual_counts.append(actual_count)
  
  template = loader.get_template('show_family_number.html')
  context = {
    'students_counts':zip(questions_students,expect_counts,actual_counts),
  }
  return HttpResponse(template.render(context,request))  


def delete_by_unique_id(unique_id):
   result = '数据已经删除'
   path = settings.MEDIA_ROOT + unique_id

   #删除数据库中的数据
   Family.objects.filter(unique_id=unique_id).delete()
   Routing.objects.filter(unique_id=unique_id).delete()
   #删除压缩文件
   if os.path.exists(path + '.zip'):
      os.remove(path+'.zip')
                       
   #删除用于判断同住人员数量的excel文件
   if os.path.exists(path + '.xlsx'):
      os.remove(path+'.xlsx')

   #删除解压缩的目录和文件
   if os.path.exists(path):
      shutil.rmtree(path)
   return result

def remove_expire_data(request):
    path = settings.MEDIA_ROOT
    files = os.listdir(path)
    hour = 36000
    for f in files:
        if os.path.exists(path + f):
          unique_id = f[0:8]
          file_ctime = os.path.getmtime(path + '/' + f)
          delta = time.time() - file_ctime 
          #删除超过十个小时的数据
          if delta > hour:
            delete_by_unique_id(unique_id)
        
    msg = '所有超过10个小时的数据都被删除了.'
    response = back_to_main_page(msg)
    return response


#删除对应的数据  
def delete_info(request):
  unique_id = get_unique_id(request)
  #用户只能查看自己上传的数据
  if unique_id is None:
    response = HttpResponse('请先上传数据!')
    return response
    
  msg = delete_by_unique_id(unique_id)
  response = back_to_main_page(msg)
  return response

def how_to_export_zip(request):
  template = loader.get_template('how_to_export_zip.html')
  context = {
              'test':'test',
            }
  return HttpResponse(template.render(context,request))
